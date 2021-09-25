from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
import pandas as pd
from pandas import ExcelWriter
test_result_location = 'test_result/Test_result.xlsx'
import os
import openpyxl
import glob
from datetime import datetime

def remove_file():
    files = glob.glob(test_result_location)
    os.remove(test_result_location)
    print("All Files have been removed")
def excel_creator():
    if(os.path.exists(test_result_location)):
        print("File found, Loading existing Excel Files")
        workbook = openpyxl.load_workbook(test_result_location)
        worksheet1 = workbook.get_sheet_by_name('Summary')
        worksheet2 = workbook.get_sheet_by_name('Details')
        return workbook,worksheet1,worksheet2
    else:
        print("File not found,creating new Excel File")
        workbook = openpyxl.Workbook()
        worksheet1 = workbook.create_sheet('Summary')
        worksheet2 = workbook.create_sheet('Details')
        return workbook,worksheet1,worksheet2
def write_header():
    #workbook = openpyxl.Workbook()
    workbook,worksheet1,worksheet2 = excel_creator()
    worksheet2.cell(row=1,column=1).value = 'S.NO'
    worksheet2.cell(row=1, column=2).value= 'Test_summary'
    worksheet2.cell(row=1,column=3).value = 'Result'
    worksheet2.cell(row=1,column=4).value = 'Remarks'
    workbook.save(test_result_location)
def write_excel(sn,test_summary,result,remarks):
    workbook,worksheet1,worksheet2 = excel_creator()
    fieldnames = (int(sn),test_summary,result,str(remarks))
    start_column = 1
    start_row = sn+ 1
    for fieldnames in fieldnames:
        worksheet2.cell(row=start_row,column=start_column).value = fieldnames
        start_column += 1
    format_excel(worksheet2,start_row)
    fit_column(worksheet2)
    workbook.save(test_result_location)
def format_excel(worksheet,start_row):
    redFill = PatternFill(start_color='EE1111',end_color='EE1111',fill_type='solid')
    greenFill = PatternFill(start_color='00AA00', end_color='00AA00', fill_type='solid')
    blueFill = PatternFill(start_color='68A0F9', end_color='68A0F9', fill_type='solid')
    character = ('A','B','C','D')
    for ranges in character:
             cell = ranges+str(start_row)
             worksheet.conditional_formatting.add('A1:D1',FormulaRule(formula=['ISBLANK(L1)'],stopIfTrue=True , fill=blueFill))
             worksheet.conditional_formatting.add(cell,FormulaRule(formula=['ISNUMBER(SEARCH("Fail",'+cell+'))'], stopIfTrue=True, fill=redFill))
             worksheet.conditional_formatting.add(cell, FormulaRule(formula=['ISNUMBER(SEARCH("PASS",' + cell + '))'],stopIfTrue=True, fill=greenFill))
def fit_column(worksheet2):
    for col in worksheet2.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value))>max_length:
                    max_length = len(cell.value)
            except:
                pass
            adjusted_width = (max_length + 0.5)
            worksheet2.column_dimensions[column].width = adjusted_width
def write_summary():
    workbook, worksheet1,worksheet2 = excel_creator()
    worksheet1.cell(row=1, column=1).value = "Test started on:"
    worksheet1.cell(row=1, column=2).value = datetime.today()
    worksheet1.cell(row=2,column=1).value = "Test Executed on:"
    worksheet1.cell(row=2, column=2).value = datetime.now()
    worksheet1.cell(row=3, column=1).value = "Total Number of Test"
    worksheet1.cell(row=3, column=2).value = "=COUNT(Details!A:A)"
    worksheet1.cell(row=4, column=1).value = "Number of Passed Test Case"
    worksheet1.cell(row=4, column=2).value = '=COUNTIF(Details!C:C,"PASS")'
    worksheet1.cell(row=5, column=1).value = "Number of Failed Test Case"
    worksheet1.cell(row=5, column=2).value = '=COUNTIF(Details!C:C,"FAIL")'
    worksheet1.cell(row=6, column=1).value = "Number of Skipped Tested Case"
    worksheet1.cell(row=6, column=2).value = '=COUNTIF(Details!D:D,"Test was Skipped due to N flag")'
    fit_column(worksheet1)
    workbook.save(test_result_location)


